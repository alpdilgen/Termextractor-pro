"""
Format exporter for TermExtractor-Pro
Exports to: XLSX, CSV, TBX, JSON
"""

import json
import csv
from pathlib import Path
from typing import List, Dict, Any, Optional
from io import BytesIO
from src.models import Term, ExtractionResult


class FormatExporter:
    """Export extraction results in various formats"""
    
    def __init__(self):
        """Initialize exporter"""
        pass
    
    def export(
        self,
        result: ExtractionResult,
        output_format: str = 'xlsx',
    ) -> bytes:
        """
        Export results in specified format.
        
        Args:
            result: ExtractionResult object
            output_format: Format ('xlsx', 'csv', 'tbx', 'json')
            
        Returns:
            Exported data as bytes
        """
        format_lower = output_format.lower().strip('.')
        
        if format_lower == 'xlsx':
            return self.export_xlsx(result)
        elif format_lower == 'csv':
            return self.export_csv(result)
        elif format_lower == 'tbx':
            return self.export_tbx(result)
        elif format_lower == 'json':
            return self.export_json(result)
        else:
            raise ValueError(f"Unsupported format: {output_format}")
    
    def export_xlsx(self, result: ExtractionResult) -> bytes:
        """Export to XLSX format"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.worksheet.table import Table, TableStyleInfo
        except ImportError:
            raise ImportError("openpyxl not installed. Install with: pip install openpyxl")
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # Sheet 1: Terms
        ws_terms = wb.create_sheet("Terms")
        self._populate_terms_sheet(ws_terms, result.terms)
        
        # Sheet 2: Derivatives (if any)
        derivatives_data = self._get_derivatives_summary(result.terms)
        if derivatives_data:
            ws_deriv = wb.create_sheet("Derivatives")
            self._populate_derivatives_sheet(ws_deriv, derivatives_data)
        
        # Sheet 3: Statistics
        ws_stats = wb.create_sheet("Statistics")
        self._populate_statistics_sheet(ws_stats, result)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    def export_csv(self, result: ExtractionResult) -> bytes:
        """Export to CSV format"""
        output = BytesIO()
        
        # Prepare data
        rows = [term.to_export_row() for term in result.terms]
        
        if not rows:
            output.write(b"No terms to export")
            return output.getvalue()
        
        # Get headers from first row
        headers = list(rows[0].keys())
        
        # Write CSV
        writer = csv.DictWriter(output, fieldnames=headers, encoding='utf-8-sig')
        writer.writeheader()
        writer.writerows(rows)
        
        return output.getvalue()
    
    def export_tbx(self, result: ExtractionResult) -> bytes:
        """Export to TBX (TermBase eXchange) format"""
        try:
            from lxml import etree
        except ImportError:
            # Fallback to xml.etree if lxml not available
            import xml.etree.ElementTree as etree
        
        # Create TBX structure
        root = etree.Element('tbx')
        root.set('type', 'TBX-CoreStructV02')
        root.set('xml:lang', result.source_language or 'en')
        
        # Header
        header = etree.SubElement(root, 'tbxHeader')
        title = etree.SubElement(header, 'title')
        title.text = "TermExtractor-Pro Export"
        
        # Body
        body = etree.SubElement(root, 'text')
        
        for term in result.terms:
            entry = etree.SubElement(body, 'termEntry')
            entry.set('id', f"term_{result.terms.index(term)}")
            
            # Descripive group (domain level)
            desc_grp = etree.SubElement(entry, 'descripGrp')
            descrip = etree.SubElement(desc_grp, 'descrip')
            descrip.set('type', 'subjectField')
            descrip.text = term.domain
            
            if term.subdomain:
                descrip_sub = etree.SubElement(desc_grp, 'descrip')
                descrip_sub.set('type', 'subSubjectField')
                descrip_sub.text = term.subdomain
            
            if term.definition:
                descrip_def = etree.SubElement(desc_grp, 'descrip')
                descrip_def.set('type', 'definition')
                descrip_def.text = term.definition
            
            # Source language term
            lang_set_src = etree.SubElement(entry, 'langSet')
            lang_set_src.set('xml:lang', result.source_language or 'en')
            
            tig_src = etree.SubElement(lang_set_src, 'tig')
            term_elem = etree.SubElement(tig_src, 'term')
            term_elem.text = term.term
            
            # POS
            term_note = etree.SubElement(tig_src, 'termNote')
            term_note.set('type', 'partOfSpeech')
            term_note.text = term.pos
            
            # Context
            if term.context:
                context_elem = etree.SubElement(tig_src, 'descrip')
                context_elem.set('type', 'context')
                context_elem.text = term.context
            
            # Target language term (if exists)
            if term.translation and result.target_language:
                lang_set_tgt = etree.SubElement(entry, 'langSet')
                lang_set_tgt.set('xml:lang', result.target_language)
                
                tig_tgt = etree.SubElement(lang_set_tgt, 'tig')
                term_tgt = etree.SubElement(tig_tgt, 'term')
                term_tgt.text = term.translation
        
        # Convert to bytes
        if hasattr(etree, 'tostring'):
            xml_bytes = etree.tostring(root, encoding='utf-8', xml_declaration=True)
        else:
            xml_str = etree.tostring(root, encoding='unicode')
            xml_bytes = f'<?xml version="1.0" encoding="UTF-8"?>{xml_str}'.encode('utf-8')
        
        return xml_bytes
    
    def export_json(self, result: ExtractionResult) -> bytes:
        """Export to JSON format"""
        data = {
            'metadata': {
                'source_language': result.source_language,
                'target_language': result.target_language,
                'domain_hierarchy': result.domain_hierarchy,
            },
            'terms': [term.to_dict() for term in result.terms],
            'statistics': result.statistics,
            'lookup_statistics': result.lookup_statistics,
            'derivative_statistics': result.derivative_statistics,
        }
        
        json_str = json.dumps(data, ensure_ascii=False, indent=2)
        return json_str.encode('utf-8')
    
    @staticmethod
    def _populate_terms_sheet(ws, terms: List[Term]) -> None:
        """Populate XLSX terms sheet"""
        # Headers
        headers = [
            'Term', 'Translation', 'From Existing', 'Source',
            'Fuzzy %', 'Derivatives', 'Domain', 'Subdomain', 'POS',
            'Definition', 'Context', 'Relevance', 'Confidence',
            'Frequency', 'Compound', 'Abbreviation', 'Variants', 'Related'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col)
            cell.value = header
            cell.font = Font(bold=True)
        
        # Data rows
        for row_idx, term in enumerate(terms, 2):
            data = [
                term.term,
                term.translation or '',
                'Yes' if term.from_existing_translation else 'No',
                term.translation_source,
                f"{term.fuzzy_match_score:.1f}%" if term.fuzzy_match_score else '',
                '; '.join(term.discovered_derivatives or []),
                term.domain,
                term.subdomain or '',
                term.pos,
                term.definition,
                term.context,
                f"{term.relevance_score:.1f}",
                f"{term.confidence_score:.1f}",
                term.frequency,
                'Yes' if term.is_compound else 'No',
                'Yes' if term.is_abbreviation else 'No',
                '; '.join(term.variants or []),
                '; '.join(term.related_terms or []),
            ]
            
            for col, value in enumerate(data, 1):
                ws.cell(row_idx, col).value = value
    
    @staticmethod
    def _populate_derivatives_sheet(ws, derivatives_data: Dict) -> None:
        """Populate XLSX derivatives sheet"""
        ws['A1'] = 'Base Term'
        ws['B1'] = 'Derivative Count'
        ws['C1'] = 'Variants'
        
        for row_idx, (base_term, variants) in enumerate(derivatives_data.items(), 2):
            ws[f'A{row_idx}'] = base_term
            ws[f'B{row_idx}'] = len(variants)
            ws[f'C{row_idx}'] = '; '.join(variants)
    
    @staticmethod
    def _populate_statistics_sheet(ws, result: ExtractionResult) -> None:
        """Populate XLSX statistics sheet"""
        ws['A1'] = 'Metric'
        ws['B1'] = 'Value'
        
        row = 2
        
        # Statistics
        for key, value in result.statistics.items():
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            row += 1
        
        # Lookup statistics
        if result.lookup_statistics:
            ws[f'A{row}'] = '--- Bilingual Lookup ---'
            row += 1
            
            for key, value in result.lookup_statistics.items():
                ws[f'A{row}'] = key
                ws[f'B{row}'] = value
                row += 1
        
        # Derivative statistics
        if result.derivative_statistics:
            ws[f'A{row}'] = '--- Derivative Discovery ---'
            row += 1
            
            for key, value in result.derivative_statistics.items():
                ws[f'A{row}'] = key
                ws[f'B{row}'] = value
                row += 1
    
    @staticmethod
    def _get_derivatives_summary(terms: List[Term]) -> Dict[str, List[str]]:
        """Get summary of derivatives"""
        summary = {}
        
        for term in terms:
            if term.discovered_derivatives:
                summary[term.term] = term.discovered_derivatives
        
        return summary
